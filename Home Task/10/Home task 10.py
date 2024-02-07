# -*- coding: utf-8 -*-

# # Мой дядя самых честных правил
# # Когда не в шутку занемог


#Задача 10-1

with open("test1.txt", 'r') as f1:

    f1_strings=f1.readlines() #читаем
    for x in f1_strings:#печатаем
        print(x.strip())

    #создаём и пишем в новом txt
    with open('test2.txt', 'w+') as f2:

        for x in f1_strings[::-1]: #обратный порядок элементов (строк) из первого файла
            x_spl = x.split() #разделяём элементы списка (строку) на подстроки
            finish_string=x_spl[::-1] #переворачиваем подстроки
            print(*finish_string, file=f2)


# Задача 10-2
import openpyxl
wb=openpyxl.load_workbook("Лист Microsoft Excel.xlsx")
ws=wb.active


names,salary=[],[] #далее через zip создадим словарь

for i in range(ws.max_row):
    for j in range(ws.max_column):
        element = ws.cell(row=i+1,column=j+1).value
        if type(element) == str: names.append(element)
        elif type(element) == int: salary.append(element)


#словарь из имен (ключей) и зарплат (значений)
dct=(dict(zip(names, salary)))

#ИТОГОВАЯ сумма всех ЗП, которую укажем в конце
itogo=sum(dct.values())

#сортировка по зарплате (значениям словаря) от макс к мин и получаем список с подкартежами: [(имя, ЗП), (имя, ЗП)..]
sorted_dct=list(sorted(dct.items(), key=lambda x: -x[1]))


wb.create_sheet("Результат")
ws_new=wb["Результат"]


for i in range(ws.max_row):
    for j in range(ws.max_column):
        add_element=ws_new.cell(row=i + 1, column=j + 1).value=sorted_dct[i][j]

#добавляем ИТОГО
ws_new.cell(row=ws_new.max_row+1, column=1).value="ИТОГО:"
ws_new.cell(row=ws_new.max_row, column=2).value=itogo


wb.save("Лист Microsoft Excel.xlsx")



#Задача 10-3
wb=openpyxl.load_workbook("Лист Microsoft Excel1.xlsx")
ws=wb.active


lst=[]
for i in range(ws.max_row):
    for j in range(2):
        element = ws.cell(row=i+1,column=j+1).value
        if type(element) == int: lst.append(element)


max_sal=max(lst)
min_sal=min(lst)
sum_sal=sum(lst)
aver_sal=int(sum(lst)/len(lst))
med_sal=int(sum(lst[1:-1])/len(lst[1:-1]))

result=[
    ['Максимальное значение', max_sal],
    ['Минимальное значение', min_sal],
    ['Сумма', sum_sal],
    ['Среднее арифметическое',aver_sal],
    ['Медиана',med_sal]]


wb.create_sheet("Итог")
ws_new=wb["Итог"]


for i in range(len(result)):
    for j in range (2):
        ws_new.cell(row=i+1, column=j+1).value=result[i][j]

wb.save("Лист Microsoft Excel1.xlsx")


