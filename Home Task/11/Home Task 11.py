# # -*- coding: utf-8 -*-
# Задача 11-1
def free_tickets(year):
    import calendar

    all_third_thursdays = []

    if 1852<=year<=2124:
        for i in range(1,13):
            first_day_month=calendar.weekday(year, i, 1)
            if first_day_month<=3: #если четверг будет на первой неделе
                weeks = calendar.monthcalendar(year,i) #вызываем список с подсписками недель, где элементы - числа недели
                third_thursday = weeks[2][3] #3-й четверг на 3-й неделе
                all_third_thursdays.append(third_thursday) #добавляём результат в общий список

            else: #если четверг небудет на первой неделе
                weeks = calendar.monthcalendar(year, i)
                third_thursday= weeks[3][3] #3-й четверг на 4-й неделе
                all_third_thursdays.append(third_thursday)
    elif year>2124: return print(f"Это слишком далекое и туманное будущее. Будет ли существовать \"Эрмитаж\" в {year}г.?")
    elif year<1852: return print(f"В {year} году \"Эрмитажа\" как музея не существовало")

    print(f"Список дат в {year} году, когда вход в Эрмитаж бесплатен:")
    for i in range(1, 13):
        print(f"{all_third_thursdays[i-1]}.{i}.{year} г.")


n=int(input("Укажите год, чтобы узнать даты, в которые вход в Эрмитаж бесплатен: "))
free_tickets(n)
print()
# #Задача 11-2

# # 1, Маяковский, Илья, "Газпром", 300000
# # 2, Булгаков, Алексей, "Яндекс", 200000
# # 3, Достоевский, Михаил, "Гугл", 100000
# # 4, Надеждина, Александра, "Гугл", 400000
# # 5, Чехова, Людмила, "Яндекс", 150000
# # 6, Маяковский, Александр, "Газпром", 100000

with open("text.txt", "r") as f:
    s=f.readlines()

    lst=[] #через цикл будем добавлять в excel по ряду
    summa=0 #считаем ЗП

    for v in s:
        v_sp = v.strip().split(", ") #превращаем строку в список
        lst.append(v_sp[:-1]+[int(v_sp[-1])]) #добавляем полученный список в общий список, зп переводим в int
        summa+=int(v_sp[4]) #считаем ИТОГО

sorted_lst=sorted(lst, key=lambda x: (x[3], x[1], x[2])) #сортируем по компании, фамилии, имени


import openpyxl
wb=openpyxl.Workbook()
ws=wb.active

ws.append(["Номер", "Фамилия", "Имя", "Компания", "Зарплата"])
for i in sorted_lst:
    ws.append(i)

ws.append(["ИТОГО"])
ws.cell(ws.max_row,ws.max_column).value=summa


wb.save("файл.xlsx")

# #Задача 11-3
def write_roman(x):
    from roman import toRoman  # скачиваем модуль roman и используем класс toRoman для перевода

    if x>=0:
        res=toRoman(int(x))
        print(f"{x}->{res}")
    else:print("Указано отрицательное значение")

number=int(input("Ввод целого арабского числа: "))

write_roman(number)

