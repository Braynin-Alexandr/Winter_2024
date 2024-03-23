# -*- coding: utf-8 -*-
from time import time
from openpyxl import load_workbook, Workbook
from collections import Counter
from pprint import pprint
from docx import Document
import matplotlib.pyplot as plt
t1=time()
wb=load_workbook('Форма.xlsx')
ws=wb.active

pupil_vpr_info={}
pupil_absent=0

for row in range(2, ws.max_row+1):
    pupil_name=ws.cell(row, 1).value
    answers = []
    for column in range(2, ws.max_column):
        value = ws.cell(row, column).value
        if value == 'отсутствовал':
            pupil_absent+=1
            break
        else:
            if value is not None: answers.append(value)
    else:
        sum_answers=sum(answers)
        pupil_vpr_info[pupil_name] = [answers, [sum_answers], [ws.cell(row, ws.max_column).value]]
wb.save('Форма.xlsx')
data=(pupil_vpr_info, pupil_absent) #вся считанная с excel информация в виде картежа
pprint(data)
print()

#оценки за 3-ю четверть
quarter_3_marks = []
quarter_3_marks_dct = {}
quarter_3_marks_good,quarter_3_marks_bad = 0, 0


for mark_quarter_3 in pupil_vpr_info.values(): #достаём оценки за 3-ю четверть из словаря и добавляем в список
    quarter_3_marks.extend(mark_quarter_3[2])


#подсчет количества 5,4,3,2, а также подсчет хороших (4,5) и плохих оценок (3,2)
mark_counter = Counter(quarter_3_marks)
for mark, count in mark_counter.most_common():
    if mark>=4: quarter_3_marks_good+=count
    else: quarter_3_marks_bad+=count
    quarter_3_marks_dct[mark]=count


#качество
quarter_3_quality=round((quarter_3_marks_good / (quarter_3_marks_good + quarter_3_marks_bad)) * 100)

#успеваемость
mark_not2=0
for mark_not2 in range(3,6):
    mark_not2+=quarter_3_marks_dct.get(mark_not2, 0)

quarter_3_performance_marks=(round((mark_not2 / len(quarter_3_marks)) * 100)) #кол-во 3,4,5 делим на общее количество оценок

#Средняя оценка по предмету
quarter_3_avarage_mark= round(sum(quarter_3_marks)/len(quarter_3_marks))


#тест впр
test_marks,test_marks_dct = [], {}
test_marks_good, test_marks_bad= 0,0


#Кол-во и % детей, повысивших и понизивших свою оценку
pupil_improved_mark, pupil_worsened_mark=0,0
#оценки за тест из баллов и сравнение оценок теста и 3-й четверит
for mark in pupil_vpr_info.values():
    result_test = sum(mark[1]) #sum - чтобы вытащить int из списка
    mark_3qr = sum(mark[2])
    if result_test<=6:
        test_marks.append(2)
        if 2>=mark_3qr: pupil_improved_mark+=1
        else: pupil_worsened_mark+=1
    elif 7<=result_test<=11:
        test_marks.append(3)
        if 3>=mark_3qr: pupil_improved_mark += 1
        else: pupil_worsened_mark += 1
    elif 12<=result_test<=15:
        test_marks.append(4)
        if 4 >= mark_3qr: pupil_improved_mark += 1
        else: pupil_worsened_mark += 1
    else:
        test_marks.append(5)
        if 5>= mark_3qr: pupil_improved_mark += 1
        else: pupil_worsened_mark+= 1

print(pupil_improved_mark)
input()

#подсчет количества 5,4,3,2, а также подсчет хороших (4,5) и плохих оценок (3,2)
mark_counter = Counter(test_marks)
for mark, count in mark_counter.most_common():
    if mark>=4: test_marks_good+=count
    else: test_marks_bad+=count
    test_marks_dct[mark]=count

#качество
test_quality=round((test_marks_good/(test_marks_good+test_marks_bad))*100)

#успеваемость
mark_not2=0
for mark_not2 in range(3,6):
    mark_not2+=test_marks_dct.get(mark_not2, 0)

test_performance_marks=(round((mark_not2 / len(test_marks)) * 100)) #кол-во 3,4,5 делим на общее количество оценок

#Средняя оценка ВПР
test_avarage_mark= round(sum(test_marks)/len(test_marks))

#Среднее кол-во решенных заданий
test_task_done=0
for i in pupil_vpr_info.values():
    for j in i[0][0:]:
        if j>0: test_task_done+=1
test_avarage_task_done=round(test_task_done/(len(pupil_vpr_info)))





#результаты подсчета
print('Количество учеников класса, которые писали ВПР:', len(pupil_vpr_info))
print('Количество учеников класса, которые не писали ВПР:', pupil_absent)
print()

print('Оценка','3-я четверть','ВПР')
i=5
while i!=2:
    print(f'{i}:',quarter_3_marks_dct.get(i,'-'), test_marks_dct.get(i,'-'))
    i-=1
else:
    print('% качества:', quarter_3_quality, test_quality)
    print('% успеваемости:', quarter_3_performance_marks, test_performance_marks)


print('Средний балл по предмету за 3-ю четверть:', quarter_3_avarage_mark)
print(f'Средний балл за ВПР: {test_avarage_mark}')
print(f'Кол-во и % детей, повысивших свою оценку: {pupil_improved_mark} ({round((pupil_improved_mark/len(pupil_vpr_info))*100)}%)')
print(f'Кол-во и % детей, понизивших свою оценку: {pupil_worsened_mark} ({round((pupil_worsened_mark/len(pupil_vpr_info))*100)}%)')


print('Проверка достоверности результатов:')

